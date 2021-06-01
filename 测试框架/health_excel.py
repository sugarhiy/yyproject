#程序媛：JY子
from openpyxl import load_workbook


class testexcel:
    def create_excel(self):
        from openpyxl import Workbook  # command+右键进入方法内，sheet是页签
        wb = Workbook()  # 工作簿
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"
        # 填入每一行中，所以存放在某个列表中
        self.height = [180, 170, 169, 158, 145]  #多个人身高
        self.weight = [66, 60, 40, 30, 100]  # 多个人体重
        # print(self.weight[0])  # 列表的索引
        # 写入要用for循环
        for i in range(len(self.height)):
            ws.cell(row=i + 2, column=1, value=self.height[i])  # 抽离变与不变的东西，
            ws.cell(row=i + 2, column=2, value=self.weight[i])  # 抽离变与不变的东西，column是列，row是行，value是写入的值
        # 最后一定要保存s
        wb.save("sample.xlsx")

    def health_weight(self):
         ld=load_workbook(filename="sample.xlsx")#读数据
         sheet=ld.active#激活sheet
         sheet["C1"]="备注" #通过名字找到页签
         for i in range(len(self.height)):
             height=sheet.cell(row=i+2 , column=1).value  # column是列，row是行，value是写入的值
             weight=sheet.cell(row=i+2 , column=2).value  # 抽离变与不变的东西，
             #获取身高对应的体重
             health_w=(height-70)*0.6
             if weight==health_w:
                 print("这是健康的体重",weight)
                 sheet.cell(row=i+2, column=3).value="健康体重"
            #最后一定要保存s
         ld.save("sample1.xlsx")

#实例化
test=testexcel()
test.create_excel()
test.health_weight()




