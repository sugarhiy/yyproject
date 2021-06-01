# from openpyxl import Workbook#command+右键进入方法内，sheet是页签
# # wb = Workbook()#工作簿
# #
# # # grab the active worksheet
# # ws = wb.active
# #
# # # Data can be assigned directly to cells
# # ws['A1'] = 42
# #
# # # Rows can also be appended
# # ws.append([1, 2, 3])
# #
# # # Python types will automatically be converted
# # import datetime
# # ws['A2'] = datetime.datetime.now()
# #
# # # Save the file
# # wb.save("sample.xlsx")


from openpyxl import Workbook#command+右键进入方法内，sheet是页签
wb = Workbook()#工作簿

ws = wb.active

ws['A1'] = "身高"
ws['B1'] = "体重"
#填入每一行中，所以存放在某个列表中
height=[180,178,169,158,145]#多个人身高
weight=[140,130,120,110,100]#多个人体重
print(weight[0])#列表的索引

#写入要用for循环
for i in range(len(height)):
    print(i)
    ws.cell(row=i+2,column=1,value=height[i])#抽离变与不变的东西，
    ws.cell(row=i+2,column=2,value=weight[i])#抽离变与不变的东西，column是列，row是行，value是写入的值



import datetime

wb.save("sample.xlsx")
